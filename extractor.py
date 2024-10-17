import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


# Make sure file is in the same directory as the script
######################################################
INPUT_FILE_NAME = 'posts.xml'
OUTPUT_FILE_NAME = 'posts.xlsx'
######################################################

# Parse the XML file
tree = ET.parse(INPUT_FILE_NAME)
root = tree.getroot()

# Prepare a list to hold the extracted data
data = []

# Define namespaces
namespaces = {
    'wp': 'http://wordpress.org/export/1.2/',
    'dc': 'http://purl.org/dc/elements/1.1/'
}

# Loop through each item in the XML
for item in root.findall('channel/item'):
    post_id = item.find('wp:post_id', namespaces).text if item.find('wp:post_id', namespaces) is not None else ''
    post_date = item.find('wp:post_date', namespaces).text if item.find('wp:post_date', namespaces) is not None else ''
    title = item.find('title').text if item.find('title') is not None else ''
    post_modified = item.find('wp:post_modified', namespaces).text if item.find('wp:post_modified', namespaces) is not None else ''
    published_date = item.find('wp:post_date_gmt', namespaces).text if item.find('wp:post_date_gmt', namespaces) is not None else ''
    published_time = item.find('wp:post_date_gmt', namespaces).text if item.find('wp:post_date_gmt', namespaces) is not None else ''
    link = item.find('link').text if item.find('link') is not None else ''
    creator = item.find('dc:creator', namespaces).text if item.find('dc:creator', namespaces) is not None else ''
    guid = item.find('guid').text if item.find('guid') is not None else ''
    slug = item.find('wp:post_name', namespaces).text if item.find('wp:post_name', namespaces) is not None else ''
    status = item.find('wp:status', namespaces).text if item.find('wp:status', namespaces) is not None else ''
    categories = [cat.text for cat in item.findall('category')] if item.findall('category') else ''
    category = ', '.join(categories)
    
    # Append the extracted data to the list
    data.append([post_id, post_date, title, post_modified, published_date, published_time, link, creator, guid, slug, status, category])

# Create a DataFrame
df = pd.DataFrame(data, columns=['post_id', 'post_date', 'title', 'post_modified', 'published_date', 'published_time', 'link', 'creator', 'guid', 'slug', 'status', 'category'])

# Sort the DataFrame by the 'post_modified' column
df = df.sort_values(by='post_modified', ascending=False)
# Check if data list is empty
if not data:
    print('No data found in the XML file')

# Write the DataFrame to an Excel file
df.to_excel(OUTPUT_FILE_NAME, index=False)

### Styling the Excel file ###

# Load the workbook and access the active sheet
wb = load_workbook(OUTPUT_FILE_NAME)
ws = wb.active

# Freeze the top row
ws.freeze_panes = 'A2'

# Set column widths
column_widths = {
    'A': 15,  # post_id
    'B': 20,  # post_date
    'C': 30,  # title
    'D': 20,  # post_modified
    'E': 20,  # published_date
    'F': 20,  # published_time
    'G': 30,  # link
    'H': 20,  # creator
    'I': 30,  # guid
    'J': 20,  # slug
    'K': 15,  # status
    'L': 30   # category
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Define the fill colors
light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")


# Iterate through the rows and apply the fill color based on the status
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    status_cell = row[10]  # Assuming the status column is the 11th column (K)
    if status_cell.value == 'draft':
        for cell in row:
            if cell.column == 11:
                cell.fill = light_gray_fill
    elif status_cell.value == 'publish':
        for cell in row:
            if cell.column == 11:
                cell.fill = light_green_fill


# Save the workbook
wb.save(OUTPUT_FILE_NAME)